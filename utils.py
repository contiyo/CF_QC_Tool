from arcgis.features import FeatureLayer
from arcgis.gis import GIS
import arcgis
from arcgis.geometry import Geometry
from arcgis.geometry.filters import intersects
import logging
import smtplib
import openpyxl
from email.message import EmailMessage
from email.contentmanager import ContentManager
import os


# ACCESSING AGOL AND GENERATING TOKEN
username = 'praveenmp'
password = 'raph1aP4'
gis = GIS("https://www.arcgis.com", username, password)

logger = logging.getLogger('__main__.' + __name__)

def new_fields(layerid):
    item = gis.content.get(layerid)
    fl_url = item.layers[0].url
    fl = FeatureLayer(fl_url, gis)
    # Field 1
    new_field = {
        "name": "poles_fail_rate",
        "type": "esriFieldTypeDouble",
        "alias": "Poles Failure Percentage (%)",
        "sqlType": "sqlTypeDouble",
        "nullable": True,
        "editable": True,
        "visible": True,
        "domain": None
    }
    update_dict = {"fields": [new_field]}
    try:
        fl.manager.add_to_definition(update_dict)
        logging.info("Field: <{}>  added in Feature Layer <{}> successfully.".format(new_field["name"], item.title))
    except Exception as e:
        logging.info('Field <{}> already existing.'.format(new_field["name"]))
        logging.debug(e)
    # Field 2
    new_field = {
        "name": "total_poles",
        "type": "esriFieldTypeInteger",
        "alias": "Total Poles Count",
        "nullable": True,
        "editable": True,
        "visible": True,
        "domain": None
    }
    update_dict = {"fields": [new_field]}
    try:
        fl.manager.add_to_definition(update_dict)
        logging.info("Field: <{}>  added in Feature Layer <{}> successfully.".format(new_field["name"], item.title))
    except Exception as e:
        logging.info('Field <{}> already existing.'.format(new_field["name"]))
        logging.debug(e)
    # Field 3
    new_field = {
        "name": "fail_poles",
        "type": "esriFieldTypeInteger",
        "alias": "QC Fail Poles",
        "nullable": True,
        "editable": True,
        "visible": True,
        "domain": None
    }
    update_dict = {"fields": [new_field]}
    try:
        fl.manager.add_to_definition(update_dict)
        logging.info("Field: <{}>  added in Feature Layer <{}> successfully.".format(new_field["name"], item.title))
    except Exception as e:
        logging.info('Field <{}> already existing.'.format(new_field["name"]))
        logging.debug(e)
    # Field 4
    new_field = {
        "name": "pass_poles",
        "type": "esriFieldTypeInteger",
        "alias": "QC Pass Poles",
        "nullable": True,
        "editable": True,
        "visible": True,
        "domain": None
    }
    update_dict = {"fields": [new_field]}
    try:
        fl.manager.add_to_definition(update_dict)
        logging.info("Field: <{}>  added in Feature Layer <{}> successfully.".format(new_field["name"], item.title))
    except Exception as e:
        logging.info('Field <{}> already existing.'.format(new_field["name"]))
        logging.debug(e)

    # Field 5
    new_field = {
        "name": "chambers_fail_rate",
        "type": "esriFieldTypeDouble",
        "alias": "Chamber Failure Percentage (%)",
        "sqlType": "sqlTypeDouble",
        "nullable": True,
        "editable": True,
        "visible": True,
        "domain": None
    }
    update_dict = {"fields": [new_field]}
    try:
        fl.manager.add_to_definition(update_dict)
        logging.info("Field: <{}>  added in Feature Layer <{}> successfully.".format(new_field["name"], item.title))
    except Exception as e:
        logging.info('Field <{}> already existing.'.format(new_field["name"]))
        logging.debug(e)
    # Field 6
    new_field = {
        "name": "total_chambers",
        "type": "esriFieldTypeInteger",
        "alias": "Total Chamber Count",
        "nullable": True,
        "editable": True,
        "visible": True,
        "domain": None
    }
    update_dict = {"fields": [new_field]}
    try:
        fl.manager.add_to_definition(update_dict)
        logging.info("Field: <{}>  added in Feature Layer <{}> successfully.".format(new_field["name"], item.title))
    except Exception as e:
        logging.info('Field <{}> already existing.'.format(new_field["name"]))
        logging.debug(e)
    # Field 7
    new_field = {
        "name": "fail_chambers",
        "type": "esriFieldTypeInteger",
        "alias": "QC Fail Chamber",
        "nullable": True,
        "editable": True,
        "visible": True,
        "domain": None
    }
    update_dict = {"fields": [new_field]}
    try:
        fl.manager.add_to_definition(update_dict)
        logging.info("Field: <{}>  added in Feature Layer <{}> successfully.".format(new_field["name"], item.title))
    except Exception as e:
        logging.info('Field <{}> already existing.'.format(new_field["name"]))
        logging.debug(e)
    # Field 8
    new_field = {
        "name": "pass_chambers",
        "type": "esriFieldTypeInteger",
        "alias": "QC Pass Chamber",
        "nullable": True,
        "editable": True,
        "visible": True,
        "domain": None
    }
    update_dict = {"fields": [new_field]}
    try:
        fl.manager.add_to_definition(update_dict)
        logging.info("Field: <{}>  added in Feature Layer <{}> successfully.".format(new_field["name"], item.title))
    except Exception as e:
        logging.info('Field <{}> already existing.'.format(new_field["name"]))
        logging.debug(e)


def spatial_selection_update(boundary_layer, boundary_feature_set, intersecting_layer, feature_type, field_to_update):
    update_features = []
    for feature in boundary_feature_set.features:
        qc_boundary_geom_dict = feature.geometry
        boundary_sr = qc_boundary_geom_dict['spatialReference']
        qc_boundary_geom = Geometry(qc_boundary_geom_dict)
        fl_filter = intersects(qc_boundary_geom, sr=boundary_sr)
        selection_layer = intersecting_layer.query(geometry_filter=fl_filter)
        count = 0
        for element in selection_layer.features:
            if feature_type == 'Pass_poles':
                if element.attributes['Type_'] == 'Poles':
                    count += 1
            if feature_type == 'Pass_chambers':
                if element.attributes['Type_'] == 'Chambers':
                    count += 1
            if feature_type == 'Fail_poles':
                if element.attributes['error_type'] == '1':
                    count += 1
            if feature_type == 'Fail_Chambers':
                if element.attributes['error_type'] == '3':
                    count += 1
        feature.attributes[field_to_update] = count
        update_features.append(feature)
    ds_flayer = boundary_layer[0]
    ds_flayer.edit_features(updates=update_features)
    logging.info('QC Polygon layer update successfully')


def get_wkid(agol_item):
    if agol_item.spatialReference is None:
        item = agol_item.layers[0]
        if item.properties['extent']['spatialReference']['wkid'] is not None:
            return item.properties['extent']['spatialReference']['wkid']
        else:
            item_url = agol_item.layers[0].url
            item_fl = arcgis.features.FeatureLayer(item_url, gis)
            item_fset = item_fl.query()
            try:
                for feature in item_fset:
                    wkid = feature.geometry['spatialReference']['wkid']
                    break
            except:
                wkid = None
                pass
            return wkid
    else:
        return agol_item.spatialReference


def send_email(to_addr, subject, body):
    from_addr = 'kpikos@entegro.ie'
    app_password = 'lqlfndqdfnbpcnrx'
    # Set up the SMTP server
    server = smtplib.SMTP('smtp.office365.com', 587)
    server.starttls()

    # Log in to your Outlook account
    server.login(from_addr, app_password)

    # Define the email message
    msg = f'From: {from_addr}\r\nTo: {to_addr}\r\nSubject: {subject}\r\n\r\n{body}'

    # Send the email
    server.sendmail(from_addr, to_addr, msg)

    # Close the SMTP connection
    server.quit()




def write_lists_to_excel(lists, sheet_names, excel_name):
    # create a new workbook
    workbook = openpyxl.Workbook()

    # loop over the lists and create a new worksheet for each list
    for i, my_list in enumerate(lists):
        # get the sheet name from the sheet_names list
        sheet_name = sheet_names[i]

        # create a new worksheet with the specified name
        worksheet = workbook.create_sheet(title=sheet_name)

        # define the headers for the worksheet
        headers = list(my_list[0].keys())

        # write the headers to the first row of the worksheet
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header

        # write the data to the worksheet
        for row_num, row_data in enumerate(my_list, 2):
            for col_num, cell_value in enumerate(row_data.values(), 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = str(cell_value)

    # delete the default 'Sheet' sheet
    default_sheet = workbook['Sheet']
    workbook.remove(default_sheet)

    # save the workbook
    workbook.save(excel_name)

def send_email2(to_addr, subject, body, file_path):
    from_addr = 'automation.account@entegro.ie'
    app_password = '@Automation12345'

    # Set up the SMTP server
    server = smtplib.SMTP('smtp.office365.com', 587)
    server.starttls()

    # Log in to your Outlook account
    server.login(from_addr, app_password)

    # Define the email message
    msg = EmailMessage()
    msg['From'] = from_addr
    if len(to_addr) == 1:
        msg['To'] = to_addr[0]
    else:
        msg['To'] = ', '.join(to_addr)
    msg['Subject'] = subject
    msg.set_content(body)

    with open(file_path, 'rb') as f:
        file_data = f.read()
    msg.add_attachment(file_data, maintype="application", subtype="xlsx", filename=file_path)

    # Send the email
    server.send_message(msg, from_addr=from_addr, to_addrs=to_addr)

    # Close the SMTP connection
    server.quit()

def delete_file(file_path):
    """Delete the file at the given path if it exists."""
    if os.path.isfile(file_path):
        os.remove(file_path)
